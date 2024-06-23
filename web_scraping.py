import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import openpyxl


pages_prefix = 'http://dnd5e.wikidot.com/'
pages_sufix = ['artificer', 'barbarian', 'bard', 'cleric', 'druid', 'fighter', 'monk', 'paladin', 'ranger', 'rogue', 'sorcerer', 'warlock', 'wizard']

pages = {}

for classe in pages_sufix:
    pages[pages_prefix + classe] = classe


def formatar_texto(texto):
    if isinstance(texto, str):
        return texto.lower().replace(' ', '-')
    return texto
    

def formata_dict(dicionario):
    return {
        formatar_texto(k): list(map(formatar_texto, v)) if isinstance(v, list) else formatar_texto(v)
        for k, v in dicionario.items()
    }


def extract_information_from_div(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    data = []

    # Extrair o nome do título da página
    div_nome = soup.find('div', class_='page-title page-header')
    if div_nome:
        span_nome = div_nome.find('span')
        if span_nome:
            data.append(('title-page', (span_nome.get_text().strip() + ' INICIO ' + '*' * 100)))

    target_div = soup.find('div', id='page-content')

    if not target_div:
        print("Div com o id 'page-content' não encontrada.")
        return []

    def tratando_div(element, data):
        if isinstance(element, str):  # Verifica se o elemento é uma string (NavigableString)
            text_content = element.strip()
            if text_content:
                data.append(('text', text_content))
        elif element.name and re.match(r'h[1-6]', element.name, re.IGNORECASE):
            current_title = element.get_text().strip() + ' TITLE '
            data.append((current_title, []))
        elif element.name == 'table':
            table_data = []
            rows = element.find_all('tr')
            for row in rows:
                cols = row.find_all(['td', 'th'])
                table_data.append([col.get_text().strip() for col in cols])
            data.append(('table', table_data))
        elif element.name == 'div':
            # Verificar se a div tem a classe 'floatright'
            if 'floatright' in element.get('class', []):
                return  # Ignorar esta div e não processá-la
            for child in element.children:
                tratando_div(child, data)
        elif element.get_text().strip():
            text_content = element.get_text().strip()
            data.append(('text', text_content))

    # Processa todos os elementos dentro de 'page-content'
    for element in target_div.children:
        tratando_div(element, data)

    return data


def cria_excel(dict_links, nome_excel):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([""])

    for url, classe in dict_links.items():
        print(f'Extraindo informações de {classe} ({url})')
        lista_informacao_pagina = extract_information_from_div(url)

        for titulo, conteudo in lista_informacao_pagina:
            if isinstance(conteudo, list):  # Conteúdo de tabela ou título com lista vazia
                ws.append([titulo])  # Adiciona o título como uma linha
                for linha in conteudo:
                    ws.append(linha)
            elif isinstance(conteudo, str):  # Conteúdo de texto
                if '\n' in conteudo:
                    for linha in conteudo.split('\n'):
                        if linha.strip():
                            ws.append([linha.strip()])
                else:
                    if conteudo.strip():
                        ws.append([conteudo.strip()])

    wb.save(f'{nome_excel}.xlsx')


def cria_dict_sub_classe():
    lista_arquetipos = ['Specialty', 'Path', 'College', 'Domain', 'Circle', 'Archetype', 'Tradition', 'Oath', 'Conclave', 'Origin', 'Patron', 'School']
    df = pd.read_excel('classes.xlsx', header=None)
    dict_sub_classes = {}
    ativador = False
    nome_classe = None

    for index, row in df.iterrows():
        celula = row[0]
        valor_coluna_ao_lado = row[1]

        if 'INICIO' in celula:
            nome_classe = celula.split()[0]

        elif 'TITLE' in celula:
            ativador = False

        elif celula in lista_arquetipos:
            ativador = True

        elif ativador and pd.notna(valor_coluna_ao_lado):
            dict_sub_classes[celula] = nome_classe

    return dict_sub_classes


def cria_link_sub_classe(dict_sub_classe):
    dict_links = {}
    pages_prefix = 'http://dnd5e.wikidot.com/'

    dict_sub_classe_formatado = formata_dict(dict_sub_classe)

    for sub_classe, classe in dict_sub_classe_formatado.items():
        dict_links[f'{pages_prefix}{classe}:{sub_classe}'] = classe


    return dict_links


def edita_excel(nome_excel):
    wb = openpyxl.load_workbook(nome_excel)
    ws = wb.active

    pattern = ["The page", "Create page"]

    rows_to_delete = []

    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
        for cell_value in row:
            if isinstance(cell_value, str) and any(pat in cell_value for pat in pattern):
                rows_to_delete.append(row[0])  # Adiciona o número da linha à lista para exclusão


    # Remover linhas marcadas para exclusão
    for row_idx_str in rows_to_delete:
        # Encontra o número da linha correspondente ao valor da célula
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            if row[0].value == row_idx_str:  # row[0] contém o valor da célula na primeira coluna da linha
                ws.delete_rows(row[0].row)
                break  # Sai do loop após excluir a linha correspondente

    wb.save(nome_excel)


def classe_e_sub_classe():
    #através do dicionário com links das classes vai criar um excel com todas as informações das páginas encontradas
    cria_excel(pages, 'classes')

    #criar dicionários com os links das sub classes através da informação encontrada sobre elas no excel de classes
    dict_sub_classes = cria_dict_sub_classe()
    dict_link_sub_classe = cria_link_sub_classe(dict_sub_classes)

    #cria o excel das sub classes da mesma maneira que o de classes
    cria_excel(dict_link_sub_classe, 'teste2')

    #tira elementos repetidos do excel de páginas que existem mas não tem nenhuma informação útil
    edita_excel('teste2.xlsx')


def itens():
    dict_items_links = {}
    dict_items_links_formatado = {}
    lista_itens = ['Adventuring Gear', 'Armor', 'Trinkets', 'Weapons', 'Firearms', 'Explosives', 'Wondrous Items', 'Currency', 'Poisons', 'Tools', 'Siege Equipment']
    
    for pages in lista_itens:
        dict_items_links[f'{pages_prefix}{pages}'] = pages

    dict_items_links_formatado = formata_dict(dict_items_links)

    cria_excel(dict_items_links_formatado, 'items')


if __name__ == "__main__":
    #classe_e_sub_classe()
    itens()
    pass


